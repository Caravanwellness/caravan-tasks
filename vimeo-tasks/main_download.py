import csv
import os
from thumbnails import get_thumbnail_url, download_thumbnail
from request import get_video_url_ott
from videos import download_video
from texttracks import download_texttracks_vimeo_ott, update_language_columns
import openpyxl
# Load environment variables

def create_folder(folder_name):
    """Create folder if it doesn't exist"""
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Created '{folder_name}' folder")


def download_thumbnail_vimeo_ott(rows):
    success_count = 0
    fail_count = 0

    for i, row in enumerate(rows):
        row_num = i + 1
        if row_num >= 3:
            continue

        video_name = row.get('Video Name', '').strip()
        url = row.get('URL', '').strip()

        if not url:
            print(f"{row_num + 1}. Skipping row - no URL")
            continue

        # print(f"{row_num}. Processing: {video_name}")
        # print(f"  URL: {url}")

        thumbnail_url = get_thumbnail_url(url)

        # video_id = url[18:28]
        # video_id = get_vimeo_url(url)
        # print(f"{row_num + 1}. Processing: {video_name} (ID: {video_id})")
    

        # Fetch texttracks

        if url:
            print(f"{row_num + 1}. Processing: {video_name} (API URL found)")
            if thumbnail_url:
                download_thumbnail(thumbnail_url, video_name, row_num)
                success_count += 1
            else:
                fail_count += 1



        # if video_api_url:
        #     print(f"  Video API URL: {video_api_url}")

        #     downloaded_languages = download_texttracks_vimeo_ott(video_api_url, video_name, row_num)

        #     if downloaded_languages:
        #         update_language_columns(row, downloaded_languages, language_columns)
        #         success_count += 1
        #     else:
        #         fail_count += 1
        # if video_id:
        #     # print(f"  Video ID: {video_id}")

        #     downloaded_languages = download_texttracks_vimeo_api(video_id, video_name, row_num)

        #     if downloaded_languages:
        #         update_language_columns(row, downloaded_languages, language_columns)
        #         success_count += 1
        #     else:
        #         fail_count += 1
        else:
            fail_count += 1

    print(f"\n{'='*60}")
    print(f"Download complete!")
    print(f"Successfully downloaded: {success_count}")
    print(f"Failed: {fail_count}")
    print(f"{'='*60}\n")


def read_csv(folder_path, file_path, endpoint, asset_to_download):
    
    # create_folder('images/thumbnails')
    # create_folder('videos')
    # create_folder('subtitles')

    print(f"\nReading URLs from: {file_path}\n")

    # Read all rows into memory so we can update and write back
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    # Identify language columns (everything between 'URL' and 'Extras')
    # url_idx = fieldnames.index('Vimeo Link')
    # extras_idx = fieldnames.index('Extras')
    # language_columns = fieldnames[url_idx + 1 : extras_idx]


    success_count = 0
    fail_count = 0

    for i, row in enumerate(rows):
        row_num = i + 1
        # if row_num >= 3:
        #     continue

        video_name = row.get('Video Name (English)', '').strip()
        url = row.get('URL', '').strip()

        if not url:
            print(f"{row_num + 1}. Skipping row - no URL")
            continue

        
        if endpoint == "vimeo_ott":
            match asset_to_download:
                case "thumbnails":
                    asset = get_thumbnail_url(url)
                case "videos":
                    asset = get_video_url_ott(url)

        if url:
            print(f"{row_num + 1}. Processing: {video_name} (API URL found)")
            if asset:

                if endpoint == "vimeo_ott":
                    match asset_to_download:
                        case "thumbnails":
                            download_thumbnail(folder_path, asset, video_name, row_num)
                        case "videos":
                            download_video(folder_path, asset, video_name, row_num)

                success_count += 1
            else:
                fail_count += 1
        else:
            fail_count += 1
    print(f"\n{'='*60}")
    print(f"Video download complete!")
    print(f"Successfully downloaded : {success_count}")
    print(f"Failed: {fail_count}")
    print(f"{'='*60}\n")

    # Write updated rows back to CSV
    # with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    #     writer = csv.DictWriter(f, fieldnames=fieldnames)
    #     writer.writeheader()
    #     writer.writerows(rows)

def read_xlsx(folder_path, file_path, endpoint, asset_to_download):
    print("\nLoading Excel file...")
    wb = openpyxl.load_workbook(file_path, read_only=False)

    # Find the English Video List sheet (case-insensitive)
    sheet = None
    for sheet_name in wb.sheetnames:
        if 'sheet1' in sheet_name.lower():
            sheet = wb[sheet_name]
            print(f"Found sheet: '{sheet_name}'")
            break

    if not sheet:
        print("ERROR: Could not find 'English Video List' sheet in Excel file")
        exit(1)

    video_title_col_idx = 4
    language_col_idx = 1
    url_col_idx = 7
    row_num = 2

    MAX_ROWS = 171  # Set to None to process all rows

    success_count = 0
    fail_count = 0

    while True:


        if row_num >= MAX_ROWS:
            print(f"Reached {MAX_ROWS} rows limit (testing mode)")
            break
        video_name = sheet.cell(row=row_num, column=video_title_col_idx).value
        language = sheet.cell(row=row_num, column=language_col_idx).value
        url = sheet.cell(row=row_num, column=url_col_idx).value

        language_folder = os.path.join(folder_path, language)
        create_folder(language_folder)

        # print(row)
        # print(vimeo_id, description)
        if not url:
            print(f"\nSkipping row {row_num}: Missing url")
            row_num += 1
            continue

        if endpoint == "vimeo_ott":
            match asset_to_download:
                case "thumbnails":
                    asset = get_thumbnail_url(url)
                case "videos":
                    asset = get_video_url_ott(url)

        if url:
            print(f"{row_num + 1}. Processing: {video_name} (API URL found)")
            if asset:

                if endpoint == "vimeo_ott":
                    match asset_to_download:
                        case "thumbnails":
                            download_thumbnail(language_folder, asset, video_name, row_num)
                        case "videos":
                            download_video(language_folder, asset, video_name, row_num)

                success_count += 1
            else:
                fail_count += 1
        else:
            fail_count += 1

        row_num += 1



def main():
    file_path = os.path.join('sheets', 'Bigyellowfish.csv')
    endpoint = "vimeo_ott" # Change to "vimeo_ott" or "vimeo_api" as needed
    asset_to_download = "thumbnails"  # Change to "videos" or "texttracks" as needed

    base = os.path.basename(file_path) # Gets "my_data_file.xlsx"
    filename, file_type = os.path.splitext(base) # Splits into "my_data_file", ".xlsx"

    folder_path = f'assets/{asset_to_download}/{filename}'

    create_folder(folder_path)

    if not os.path.exists(file_path):
        print(f"Error: {file_type} file not found at {file_path}")
        return
    
    if file_type == ".xlsx":
        read_xlsx(folder_path, file_path, endpoint, asset_to_download)
        return
    else:
        read_csv(folder_path, file_path, endpoint, asset_to_download)
        return




if __name__ == '__main__':
    main()
