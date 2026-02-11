from openai import OpenAI
import anthropic
import os
import openpyxl
import difflib
from dotenv import load_dotenv
import shutil
import json
from datetime import datetime
import tempfile


def load_all_tags(csv_path):
    """Load all tags from CSV and return normalized + original versions"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        tags = [line.strip() for line in f if line.strip()]

    # Create lookup: normalized -> original
    tag_lookup = {tag.lower().strip(): tag for tag in tags}
    return tag_lookup


def validate_and_match_tags(response_tags, tag_lookup):
    """Validate tags and find closest matches using fuzzy matching"""
    matched_tags = []
    logs = []

    for tag in response_tags:
        tag_normalized = tag.strip().lower()

        # Try exact match first
        if len(matched_tags) >= 10:
            continue
        elif tag_normalized in tag_lookup:
            matched_tags.append(tag_lookup[tag_normalized])
            # logs.append(f"✓ Exact match: '{tag}' -> '{tag_lookup[tag_normalized]}'")
        else:
            # Fuzzy match
            all_tags = list(tag_lookup.keys())
            matches = difflib.get_close_matches(tag_normalized, all_tags, n=1, cutoff=0.95)

            if matches:
                matched_tag = tag_lookup[matches[0]]
                matched_tags.append(matched_tag)
                logs.append(f"⚠ Fuzzy match: '{tag}' -> '{matched_tag}' (similarity check)")
            else:
                logs.append(f"✗ No match found for: '{tag}' (skipped)")

    return matched_tags, logs


def read_csv_file(csv_path):
    """Read CSV file and return all tags as a formatted string"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        tags = [line.strip() for line in f if line.strip()]

    csv_content = f"CSV File: {csv_path}\n"
    csv_content += f"Total available tags: {len(tags)}\n\n"
    csv_content += "Complete tag list (choose ONLY from these):\n"
    csv_content += ", ".join(tags)

    return csv_content

def query_llm_with_video_and_csv(
    query,
    csv_path,
    video_info,
    model_provider="anthropic",
    model_name=None,
    api_key=None
):
    """
    Send a query to an LLM (Claude or OpenAI) with CSV data and video information

    Args:
        query (str): The question/prompt to ask the LLM
        csv_path (str): Path to the CSV file
        video_info (dict): Dictionary containing video metadata
        model_provider (str): Either "anthropic" or "openai" (default: "anthropic")
        model_name (str): Specific model to use (optional, uses defaults if not provided)
        api_key (str): API key (optional, will use env var if not provided)

    Returns:
        str: LLM's response
    """
    # Read CSV content
    csv_content = read_csv_file(csv_path)

    # Format video information
    video_content = "Video Information:\n"
    for key, value in video_info.items():
        video_content += f"- {key}: {value}\n"

    # Construct the full message
    full_message = f"""
{query}

{csv_content}

{video_content}
"""

    if model_provider.lower() == "openai":
        # Initialize OpenAI client
        if api_key is None:
            api_key = os.getenv('OPENAI_API_KEY')

        client = OpenAI(api_key=api_key)

        # Set default model if not provided
        if model_name is None:
            model_name = "gpt-4o"

        # Make the API request (OpenAI format)
        response = client.chat.completions.create(
            model=model_name,
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": full_message
                }
            ]
        )

        return response.choices[0].message.content

    elif model_provider.lower() == "anthropic":
        # Initialize Anthropic client
        if api_key is None:
            api_key = os.getenv('ANTHROPIC_API_KEY')

        client = anthropic.Anthropic(api_key=api_key)

        # Set default model if not provided
        if model_name is None:
            model_name = "claude-sonnet-4-5-20250929"

        # Make the API request (Anthropic format)
        message = client.messages.create(
            model=model_name,
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": full_message
                }
            ]
        )

        return message.content[0].text

    else:
        raise ValueError(f"Unsupported model_provider: {model_provider}. Use 'anthropic' or 'openai'.")


def create_backup(file_path):
    """Create a timestamped backup of the file"""
    if not os.path.exists(file_path):
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = file_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
    shutil.copy2(file_path, backup_path)
    print(f"✓ Backup created: {backup_path}")
    return backup_path


def save_workbook_safely(workbook, file_path):
    """Save workbook using atomic write (temp file + rename)"""
    # Create a temporary file in the same directory
    temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', dir=os.path.dirname(file_path))
    os.close(temp_fd)  # Close the file descriptor

    try:
        # Save to temp file
        workbook.save(temp_path)

        # If save succeeded, replace original file
        if os.path.exists(file_path):
            os.remove(file_path)
        shutil.move(temp_path, file_path)

        return True
    except Exception as e:
        # Clean up temp file if it exists
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise e


def load_progress(progress_file):
    """Load progress from JSON file"""
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as f:
            return json.load(f)
    return {"last_processed_row": 1, "completed_rows": []}


def save_progress(progress_file, row_number, completed_rows):
    """Save progress to JSON file"""
    progress = {
        "last_processed_row": row_number,
        "completed_rows": completed_rows,
        "last_update": datetime.now().isoformat()
    }
    with open(progress_file, 'w') as f:
        json.dump(progress, f, indent=2)


# Example usage
if __name__ == "__main__":
    load_dotenv()

    excel_file = 'assets/Caravan English Video RAG List.xlsx'
    progress_file = 'assets/progress.json'

    # Create backup before starting
    print("\n=== Creating backup ===")
    backup_file = create_backup(excel_file)

    # Load progress
    # print("\n=== Loading progress ===")
    # progress = load_progress(progress_file)
    # print(f"Last processed row: {progress['last_processed_row']}")
    # print(f"Total completed rows: {len(progress['completed_rows'])}")

    print("\n=== Loading Excel file ===")
    wb = openpyxl.load_workbook(excel_file, read_only=False)

    # Find the English Video List sheet (case-insensitive)
    english_sheet = None
    for sheet_name in wb.sheetnames:
        if 'english video list' in sheet_name.lower():
            english_sheet = wb[sheet_name]
            print(f"Found sheet: '{sheet_name}'")
            break


    if not english_sheet:
        print("ERROR: Could not find 'English Video List' sheet in Excel file")
        exit(1)


    category_col_idx = 3
    video_name_col_idx = 4
    video_length_col_idx = 5
    teacher_col_idx = 6
    description_col_idx = 15
    transcript_col_idx = 27 # Column U is index 21
    first_tag_col_idx = 28 # Column V is index 22
    row_counter = 2
    # row_counter = max(2, progress['last_processed_row'])  # Resume from last progress

    MAX_ROWS = 1720  # Set to None to process all rows

    # Initialize counters
    successful = 0
    failed = 0
    skipped = 0
    # completed_rows = progress['completed_rows'].copy()

    # CSV file path
    csv_path = "assets/Tags.csv"

    # Load tag lookup once before the loop
    tag_lookup = load_all_tags(csv_path)
    print(f"Loaded {len(tag_lookup)} tags from CSV\n")

    # Save interval (save every N successful updates)
    SAVE_INTERVAL = 5
    updates_since_save = 0

    while True:
        status = 0
        status_info = "default"

        if MAX_ROWS and row_counter >= MAX_ROWS:
            print(f"Reached {MAX_ROWS} rows limit (testing mode)")
            break

        category = english_sheet.cell(row=row_counter, column=category_col_idx).value
        video_name = english_sheet.cell(row=row_counter, column=video_name_col_idx).value
        video_length = english_sheet.cell(row=row_counter, column=video_length_col_idx).value
        teacher = english_sheet.cell(row=row_counter, column=teacher_col_idx).value
        description = english_sheet.cell(row=row_counter, column=description_col_idx).value
        transcripts = english_sheet.cell(row=row_counter, column=transcript_col_idx).value
        tag = english_sheet.cell(row=row_counter, column=first_tag_col_idx).value



        if not transcripts:
            # print(f"\nSkipping row {row_counter}: no transcript available")
            skipped += 1
            row_counter += 1
            continue
        elif tag:
            # print(f"\nSkipping row {row_counter}: tags already exist")
            skipped += 1
            # completed_rows.append(row_counter)  # Track as completed
            row_counter += 1
            continue
        else:
            # Video information
            video_info = {
                "category": category,
                "video_name": video_name,
                "description": description,
                "transcripts": transcripts
            }
            
            # Query
            query = (
                "Analyze the video transcript and description to identify its PRIMARY themes and main focus areas. "
                "Pick the 10 most relevant unique tags from the CSV that represent CENTRAL themes of the video. "
                "You may provide less than 10 tags if the next most relevant tag is too unrelated to the content of the video. "
                "\n"
                "CRITICAL TAG SELECTION CRITERIA:\n"
                "1. Each tag must represent a MAIN focus or substantial theme in the video, not just a passing mention\n"
                "2. If a concept is only mentioned briefly or in passing (1-2 times), do NOT select it as a tag\n"
                "3. The video should spend meaningful time discussing or demonstrating the tagged concept\n"
                "4. Ask yourself: 'Is this tag a primary reason someone would watch this video?' If no, don't use it\n"
                "5. Prioritize tags that match the video's category and intended learning outcomes\n"
                "\n"
                "FORMATTING REQUIREMENTS:\n"
                "- Tags MUST be exclusively chosen from and EXACTLY as written in the provided CSV tag list\n"
                "- Copy each tag character-for-character from the CSV - do not create, modify, or paraphrase any tags\n"
                "- Double-check that every tag you select appears in the CSV list above\n"
                "- Tags should be unique and should not include any duplicates\n"
                "- Format: Provide tags separated by commas without spaces after commas (tag1,tag2,tag3)\n"
                "- Provide ONLY the tags in your response, no other text or formatting"
                "- If you cannot find any relevant tags, return 'None'"
            ) \
            
            # Make the request
            try:
                

                highest_tag_count = 0
                retry_count = 1
                max_retries = 3
                saved_matched_tags = []
                saved_validation_logs = []

                while retry_count <= max_retries and highest_tag_count < 5:
                    response = query_llm_with_video_and_csv(
                        query=query,
                        csv_path=csv_path,
                        video_info=video_info,
                        model_provider="anthropic",  # Change to "openai" to use OpenAI models
                        # model_name="gpt-4o"  # Optionally specify a different model
                    )
                    print(f"\n{row_counter}: {video_name}")
                    print(f"Attempt {retry_count}; Raw response: {response}")

                    # Validate and match tags
                    response_tags = response.strip().split(',')
                    matched_tags, validation_logs = validate_and_match_tags(response_tags, tag_lookup)
                    if len(matched_tags) > highest_tag_count:
                        highest_tag_count = len(matched_tags)
                        saved_matched_tags = matched_tags
                        saved_validation_logs = validation_logs
                    retry_count += 1
                    if matched_tags[0] == 'None':
                        print(f"Received 'None' as response")
                        retry_count += 10
                    print(f"Matched tags: {matched_tags} (Count: {len(matched_tags)})")

                # Print validation logs
                for log in validation_logs:
                    print(log)

                # Write matched tags to Excel
                for index, tag in enumerate(matched_tags):
                    tag_cell = english_sheet.cell(row=row_counter, column=first_tag_col_idx + index)
                    tag_cell.value = tag

                print(f"✓ Wrote {len(matched_tags)} tags to Excel")
                successful += 1
                updates_since_save += 1
                # completed_rows.append(row_counter)

                # Save periodically using atomic write
                if updates_since_save >= SAVE_INTERVAL:
                    try:
                        print(f"\n⏳ Saving workbook (atomic write)...")
                        save_workbook_safely(wb, excel_file)
                        print(f"✓ Workbook saved successfully")

                        # Save progress
                        # save_progress(progress_file, row_counter, completed_rows)
                        # print(f"✓ Progress saved")

                        updates_since_save = 0
                    except Exception as save_error:
                        print(f"❌ Save failed: {save_error}")
                        print(f"⚠ Progress not saved. Will retry on next interval.")

            except KeyboardInterrupt:
                print("\n\n⚠ Interrupted by user. Saving progress...")
                try:
                    save_workbook_safely(wb, excel_file)
                    # save_progress(progress_file, row_counter - 1, completed_rows)
                    # print("✓ Progress saved. You can resume later.")
                except Exception as e:
                    print(f"❌ Failed to save progress: {e}")
                raise

            except Exception as e:
                print(f"❌ Error: {e}")
                failed += 1

        row_counter += 1

    # Final save at the end
    print("\n\n=== Final Save ===")
    try:
        if updates_since_save > 0:
            print("⏳ Saving final changes...")
            save_workbook_safely(wb, excel_file)
            # save_progress(progress_file, row_counter, completed_rows)
            print("✓ Final save completed")
        else:
            print("No unsaved changes")
    except Exception as e:
        print(f"❌ Final save failed: {e}")

    # Print summary
    print("\n\n=== Summary ===")
    print(f"Successful: {successful}")
    print(f"Failed: {failed}")
    print(f"Skipped: {skipped}")
    # print(f"Total completed rows: {len(completed_rows)}")
    if backup_file:
        print(f"\nBackup file: {backup_file}")
    print(f"Progress file: {progress_file}")