from dotenv import load_dotenv
from download_texttracks import get_video_id_from_uri, fetch_texttracks, select_texttrack, sanitize_filename, get_vtt, parse_vtt_to_text 
import vimeo
import os
import openpyxl
import re
import time
import csv



# Load environment variables from .env file
load_dotenv()

# Get API key from environment variable
access_token = os.getenv('VIMEO_UPLOAD_API_KEY')
client_id = os.getenv('VIMEO_CLIENT_ID')
client_secret = os.getenv('VIMEO_CLIENT_SECRET')

excel_file = 'assets/Caravan English Videos List (1-7-2026) Tags.xlsx'
log_file = 'sheets/upload_log.csv'


print("\nLoading Excel file...")
wb = openpyxl.load_workbook(excel_file, read_only=False)

# Find the English Video List sheet (case-insensitive)
english_sheet = None
for sheet_name in wb.sheetnames:
    if 'english video list' in sheet_name.lower():
        english_sheet = wb[sheet_name]
        print(f"Found sheet: '{sheet_name}'")
        break

if not english_sheet:
    print("ERROR: Could not find 'English Video List' sheet in Excel file")
    exit(1)

results = []
description_col_idx = 9
vimeo_id_col_idx = 20
vimeo_name_col_idx = 4
row_counter = 1000
transcript_col_idx = 21 # Column U is index 21

MAX_ROWS = 1834  # Set to None to process all rows

# Initialize counters
successful = 0
failed = 0
skipped = 0
series_name = ""


while True:
    status = 0
    status_info = "default"

    if MAX_ROWS and row_counter >= MAX_ROWS:
        print(f"Reached {MAX_ROWS} rows limit (testing mode)")
        break

    teacher = english_sheet.cell(row=row_counter, column=7).value
    series = english_sheet.cell(row=row_counter, column=5).value
    vimeo_id = english_sheet.cell(row=row_counter, column=vimeo_id_col_idx).value
    description = english_sheet.cell(row=row_counter, column=description_col_idx).value
    transcripts = english_sheet.cell(row=row_counter, column=transcript_col_idx).value
    vimeo_name = english_sheet.cell(row=row_counter, column=vimeo_name_col_idx).value

    # Check if we've reached the end of the data
    # if not vimeo_id and not description and not transcripts:
    #     print(f"\nReached end of data at row {row_counter}")
    #     break
    # print(row)
    # print(vimeo_id, description)
    if not teacher and vimeo_name:
        series_name = vimeo_name.strip().replace(" Series", "")
        print(f"Series name updated to: '{series_name}'")
        # print(f"\nSkipping row {row_counter}: Missing Vimeo ID")
    # elif transcripts:
    #     # print(f"\nSkipping row {row_counter}: Transcript already exists")
    #     skipped += 1
    #     row_counter += 1
    #     continue
    else:

        series_cell = english_sheet.cell(row=row_counter, column=5)
        series_cell.value = series_name
        wb.save(excel_file)
        print(f" Added series name {series_name} to Excel (Row {row_counter}, {vimeo_name})")
        successful += 1

    row_counter += 1

# Write results to log.csv
# print("\nWriting results to log.csv...")
# with open(log_file, 'w', newline='', encoding='utf-8-sig') as f:
#     fieldnames = ['title', 'vimeo_id', 'status', 'status_info', 'updated_row','description']
#     writer = csv.DictWriter(f, fieldnames=fieldnames)

#     writer.writeheader()
#     writer.writerows(results)

print(f"\n{'='*60}")
print(f"Processing complete!")
print(f"Successfully processed: {successful}")
print(f"Failed: {failed}")
print(f"Skipped: {skipped}")
print(f"Total rows processed: {row_counter - 2}")
print(f"{'='*60}")