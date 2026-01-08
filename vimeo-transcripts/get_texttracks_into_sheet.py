from dotenv import load_dotenv
from download_texttracks import get_video_id_from_uri, fetch_texttracks, select_texttrack, sanitize_filename, get_vtt, parse_vtt_to_text 
import vimeo
import os
import openpyxl
import re
import time
import csv



# Load environment variables from .env file
load_dotenv()

# Get API key from environment variable
access_token = os.getenv('VIMEO_UPLOAD_API_KEY')
client_id = os.getenv('VIMEO_CLIENT_ID')
client_secret = os.getenv('VIMEO_CLIENT_SECRET')

excel_file = 'assets/Caravan English Videos List (1-7-2026).xlsx'
log_file = 'sheets/upload_log.csv'


print("\nLoading Excel file...")
wb = openpyxl.load_workbook(excel_file, read_only=False)

# Find the English Video List sheet (case-insensitive)
english_sheet = None
for sheet_name in wb.sheetnames:
    if 'english video list' in sheet_name.lower():
        english_sheet = wb[sheet_name]
        print(f"Found sheet: '{sheet_name}'")
        break

if not english_sheet:
    print("ERROR: Could not find 'English Video List' sheet in Excel file")
    exit(1)

results = []
description_col_idx = 9
vimeo_id_col_idx = 20
row_counter = 2
transcript_col_idx = 21 # Column U is index 21

MAX_ROWS = 2000  # Set to None to process all rows

# Initialize counters
successful = 0
failed = 0
skipped = 0



while True:
    status = 0
    status_info = "default"

    if MAX_ROWS and row_counter >= MAX_ROWS:
        print(f"Reached {MAX_ROWS} rows limit (testing mode)")
        break

    vimeo_id = english_sheet.cell(row=row_counter, column=vimeo_id_col_idx).value
    description = english_sheet.cell(row=row_counter, column=description_col_idx).value
    transcripts = english_sheet.cell(row=row_counter, column=transcript_col_idx).value

    # Check if we've reached the end of the data
    # if not vimeo_id and not description and not transcripts:
    #     print(f"\nReached end of data at row {row_counter}")
    #     break
    # print(row)
    # print(vimeo_id, description)
    if not vimeo_id:
        # print(f"\nSkipping row {row_counter}: Missing Vimeo ID")
        skipped += 1
        row_counter += 1
        continue
    elif transcripts:
        # print(f"\nSkipping row {row_counter}: Transcript already exists")
        skipped += 1
        row_counter += 1
        continue
    else:
        print(f"\nProcessing row {row_counter}")
        match1 = re.search(r'vimeo\.com/(\d+)', vimeo_id if vimeo_id else '')
        video_id = match1.group(1) if match1 else None


    

        # Fetch texttracks
        texttracks = fetch_texttracks(video_id)

        if texttracks is None:
            print(f"  Failed to fetch texttracks")
            failed += 1
            row_counter += 1
            continue

        if not texttracks:
            print(f"  No texttracks available")
            row_counter += 1
            failed += 1
            continue

        # Select appropriate texttrack
        autogen, selected_track = select_texttrack(texttracks)

        # Check if transcript already exists

        # Download the VTT file
        link = selected_track.get("link")
        language = selected_track.get("language")

        if not link:
            print(f"  No download link available")
            failed += 1
            continue


        print(f"  Downloading texttrack (language: {language})")

        try:
            texttrack = get_vtt(link)
            if not texttrack:
                print(f"  Failed to download texttrack")
                failed += 1
                row_counter += 1
                continue

            print(f"  Downloaded {len(texttrack)} characters")
            # print(texttrack[:200] + "..." if texttrack and len(texttrack) > 200 else texttrack)

            transcript = parse_vtt_to_text(texttrack)
            print(f"  Parsed transcript: {len(transcript)} characters")

            transcript_cell = english_sheet.cell(row=row_counter, column=transcript_col_idx)
            transcript_cell.value = transcript
            wb.save(excel_file)
            print(f"  Saved transcript to Excel (Row {row_counter}, Column {transcript_col_idx})")
            successful += 1
            time.sleep(10)
        except Exception as e:
            print(f"  Error processing texttrack: {e}")
            failed += 1

    row_counter += 1

# Write results to log.csv
# print("\nWriting results to log.csv...")
# with open(log_file, 'w', newline='', encoding='utf-8-sig') as f:
#     fieldnames = ['title', 'vimeo_id', 'status', 'status_info', 'updated_row','description']
#     writer = csv.DictWriter(f, fieldnames=fieldnames)

#     writer.writeheader()
#     writer.writerows(results)

print(f"\n{'='*60}")
print(f"Processing complete!")
print(f"Successfully processed: {successful}")
print(f"Failed: {failed}")
print(f"Skipped: {skipped}")
print(f"Total rows processed: {row_counter - 2}")
print(f"{'='*60}")