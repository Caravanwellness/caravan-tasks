import csv
import openpyxl
import re

# File paths
csv_file = 'sheets/videos.csv'
excel_file = 'sheets/Caravan Wellness Master Video List - INTERNAL.xlsx'
log_file = 'sheets/log.csv'

def extract_vimeo_id(vimeo_url):
    """Extract the numeric Vimeo ID from a Vimeo URL."""
    if not vimeo_url:
        return None

    # Match pattern: vimeo.com/{numeric_id}/...
    match = re.search(r'vimeo\.com/(\d+)', str(vimeo_url))
    if match:
        return match.group(1)
    return None

print("Starting video matching process...")
print("=" * 60)

# Load the Excel workbook and find the English Video List sheet
print("\nLoading Excel file...")
wb = openpyxl.load_workbook(excel_file, read_only=False)

# Find the English Video List sheet (case-insensitive)
english_sheet = None
for sheet_name in wb.sheetnames:
    if 'english video list' in sheet_name.lower():
        english_sheet = wb[sheet_name]
        print(f"Found sheet: '{sheet_name}'")
        break

if not english_sheet:
    print("ERROR: Could not find 'English Video List' sheet in Excel file")
    exit(1)

# Read all video page links from the Excel sheet
print("Reading Excel data...")
excel_links = {}
headers = None
description_col_idx = None

for idx, row in enumerate(english_sheet.iter_rows(values_only=False)):
    if idx == 0:
        # Store headers to find the correct column
        headers = [str(cell.value).lower() if cell.value else '' for cell in row]
        print(f"Excel headers: {[cell.value for cell in row]}")

        # Find the description column index
        for i, header in enumerate(headers):
            if 'description' in header:
                description_col_idx = i
                print(f"Found description column at index {description_col_idx}")
                break
        continue

    # Find the column that contains video page links and vimeo links
    video_page_link = None
    vimeo_link = None
    row_number = idx + 1  # Excel row number (1-indexed, accounting for header)

    for col_idx, cell in enumerate(row):
        if cell.value and isinstance(cell.value, str):
            # Find the video page link
            if 'app.allinonewellbeing.com/videos' in cell.value:
                video_page_link = cell.value.strip()
            # Find the vimeo link
            elif 'vimeo.com' in cell.value:
                vimeo_link = cell.value.strip()

    if video_page_link:
        excel_links[video_page_link] = {
            'row_number': row_number,
            'vimeo_link': vimeo_link,
            'vimeo_id': extract_vimeo_id(vimeo_link),
        }

print(f"Found {len(excel_links)} video links in Excel file")

# Read videos.csv and match with Excel data
print("\nProcessing videos.csv (first 10 rows only for testing)...")
results = []
match_count = 0
no_match_count = 0
row_counter = 0
MAX_ROWS = 2

with open(csv_file, 'r', encoding='utf-8-sig') as f:
    csv_reader = csv.DictReader(f)

    for row in csv_reader:
        if row_counter >= MAX_ROWS:
            print(f"Reached {MAX_ROWS} rows limit (testing mode)")
            break

        title = row.get('Title', '')
        description = row.get('Description', '')
        video_page_link = row.get('Video Page Link', '').strip()
        video_link = row.get('Video Link', '').strip()

        # Check if this video page link exists in Excel
        if video_page_link in excel_links:
            excel_data = excel_links[video_page_link]
            excel_vimeo_link = excel_data.get('vimeo_link', '')
            excel_vimeo_id = excel_data.get('vimeo_id', '')
            excel_row_number = excel_data.get('row_number', '')

            # Write description to Excel if description column exists
            if description_col_idx is not None and description:
                # Write to the cell (row_number is already 1-indexed for Excel)
                cell = english_sheet.cell(row=excel_row_number, column=description_col_idx + 1)
                cell.value = description
                updated_row = excel_row_number
            else:
                updated_row = ''

            # Check if vimeo link is empty
            if not excel_vimeo_link:
                status = "no vimeo link"
            else:
                status = "matching entry found"
            match_count += 1
        else:
            status = "matching entry not found"
            no_match_count += 1
            excel_vimeo_link = ''
            excel_vimeo_id = ''
            excel_row_number = ''
            updated_row = ''

        results.append({
            'title': title,
            'url': video_page_link,
            'video_link': video_link,
            'excel_vimeo_link': excel_vimeo_link,
            'excel_vimeo_id': excel_vimeo_id,
            'status': status,
            'updated_row': updated_row
        })

        row_counter += 1

# Save the updated Excel file
print("\nSaving updated Excel file...")
wb.save(excel_file)
print(f"Excel file updated: {excel_file}")

# Write results to log.csv
print("\nWriting results to log.csv...")
with open(log_file, 'w', newline='', encoding='utf-8-sig') as f:
    fieldnames = ['title', 'url', 'video_link', 'excel_vimeo_link', 'excel_vimeo_id', 'status', 'updated_row']
    writer = csv.DictWriter(f, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(results)

# Count how many rows were updated
updated_count = sum(1 for r in results if r['updated_row'])

print(f"\n{'=' * 60}")
print(f"RESULTS:")
print(f"  Total videos processed: {len(results)}")
print(f"  Matching entries found: {match_count}")
print(f"  Matching entries not found: {no_match_count}")
print(f"  Descriptions written to Excel: {updated_count}")
print(f"\nLog file saved: {log_file}")
print(f"{'=' * 60}")
