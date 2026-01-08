from dotenv import load_dotenv
import vimeo
import os
import openpyxl
import re
import time
import csv



# Load environment variables from .env file
load_dotenv()

# Get API key from environment variable
access_token = os.getenv('VIMEO_UPLOAD_API_KEY')
client_id = os.getenv('VIMEO_CLIENT_ID')
client_secret = os.getenv('VIMEO_CLIENT_SECRET')

excel_file = 'sheets/Caravan Wellness Master Video List - INTERNAL.xlsx'
log_file = 'sheets/upload_log.csv'


print("\nLoading Excel file...")
wb = openpyxl.load_workbook(excel_file, read_only=False)

# Find the English Video List sheet (case-insensitive)
english_sheet = None
for sheet_name in wb.sheetnames:
    if 'english video list' in sheet_name.lower():
        english_sheet = wb[sheet_name]
        print(f"Found sheet: '{sheet_name}'")
        break

if not english_sheet:
    print("ERROR: Could not find 'English Video List' sheet in Excel file")
    exit(1)

results = []
description_col_idx = 9
vimeo_id_col_idx = 20
row_counter = 2
completed_col_idx = 22

MAX_ROWS = 600  # Set to None to process all rows



while True:
    status = 0
    status_info = "default"

    if row_counter >= MAX_ROWS:
        print(f"Reached {MAX_ROWS} rows limit (testing mode)")
        break
    vimeo_id = english_sheet.cell(row=row_counter, column=vimeo_id_col_idx).value
    description = english_sheet.cell(row=row_counter, column=description_col_idx).value
    completed = english_sheet.cell(row=row_counter, column=completed_col_idx).value
    # print(row)
    # print(vimeo_id, description)
    if not vimeo_id or not description:
        print(f"\nSkipping row {row_counter}: Missing Vimeo ID or Description")
        row_counter += 1
        continue
    # elif completed and "Description Updated" in str(completed):
    #     print(f"\nSkipping row {row_counter}: Description already updated")
    #     row_counter += 1
    #     continue
    else:
        match1 = re.search(r'vimeo\.com/(\d+)', vimeo_id if vimeo_id else '')
        uri = '/videos/' + match1.group(1) if match1 else None
        client = vimeo.VimeoClient(
            token=access_token,
            key=client_id,
            secret=client_secret
        )

        print(f"Attempting to update video: {uri}")
        print(f"{'='*60}")

        try:
            response = client.patch(uri, data={
                'description': description
            })

            # Check response status
            print(f"\n✓ SUCCESS!")
            print(f"Status Code: {response.status_code}")
            status = response.status_code

            success_cell = english_sheet.cell(row=row_counter, column=completed_col_idx)  # Column U is index 21
            success_cell.value = "Description Updated"
            wb.save(excel_file)



            if response.status_code == 200:
                status_info = "Video description updated successfully!"
                print("Video description updated successfully!")


                # Try to parse response body
                try:
                    response_data = response.json()
                    print(f"\nResponse Data:")
                    print(f"  - Video Name: {response_data.get('name', 'N/A')}")
                    print(f"  - Video URI: {response_data.get('uri', 'N/A')}")
                    print(f"  - Description: {response_data.get('description', 'N/A')[:100]}...")
                except:
                    print(f"\nRaw Response: {response.text[:200]}")
            elif response.status_code == 204:
                status_info = "Video updated successfully (No content returned)"
                print("Video updated successfully (No content returned)")
            else:
                print(f"Unexpected status code: {response.status_code}")
                status_info = f"Response: {response.text}"
                print(f"Response: {response.text}")
            time.sleep(5)

        except Exception as e:
            print(f"\n✗ FAILED!")
            print(f"Error Type: {type(e).__name__}")
            print(f"Error Message: {str(e)}")

            # Check if it's an HTTP error with response details
            if hasattr(e, 'response'):
                status= e.response.status_code
                status_info = e.response.text
                print(f"\nHTTP Status Code: {e.response.status_code}")
                print(f"Response Body: {e.response.text}")

            time.sleep(30)

    results.append({
        'title': english_sheet.cell(row=row_counter, column=4).value,
        'vimeo_id': vimeo_id,
        'status': status,
        'status_info' : status_info,
        'updated_row': row_counter,
        'description': description[:100] if description else ''
    })
    row_counter += 1

# Write results to log.csv
print("\nWriting results to log.csv...")
with open(log_file, 'w', newline='', encoding='utf-8-sig') as f:
    fieldnames = ['title', 'vimeo_id', 'status', 'status_info', 'updated_row','description']
    writer = csv.DictWriter(f, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(results)

print(f"\n{'='*60}")